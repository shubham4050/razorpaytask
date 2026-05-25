#!/usr/bin/env python3
"""
Shree Leepi → English (Roman) Converter for Marathi Names and Mobile Numbers

Handles 50,000+ records efficiently using conversion caching.

Usage:
  python shreelipi_to_english.py <input.xlsx> <output.xlsx>
  python shreelipi_to_english.py               # uses defaults
"""

import pandas as pd
import re, os, sys, time
from indic_transliteration import sanscript
from indic_transliteration.sanscript import transliterate

# ─────────────────────────────────────────────────────────
# SHREE LEEPI → UNICODE DEVANAGARI
# ─────────────────────────────────────────────────────────

DIGRAPHS = [
    ("m¡","ौ"),  ("mo","ो"),   ("qe","शिं"),
    ("qg","सिं"),("Am","आ"),
    ("Q>","ट"),  ("S>","ड"),   ("R>","ठ"),  ("T>","ढ"),
]

CHAR_MAP = {
    "H":"क","I":"ख","J":"ग","K":"घ","L":"ङ",
    "M":"च","N":"छ","O":"ज","P":"झ",
    "Q":"ट","R":"ठ","S":"ड","T":"ढ","U":"ण",
    "V":"त","W":"थ","X":"द","Y":"ध","Z":"न",
    "_":"म","n":"प","~":"ब","^":"भ",
    "a":"र","b":"ल","d":"व","e":"श","f":"ष",
    "g":"स","h":"ह","i":"ळ","j":"य","k":"ळ",
    "l":"श्र","p":"प्","v":"व","z":"ज","\\":"फ",
    "m":"ा","r":"ी","u":"ी","w":"ु","y":"ू",
    "o":"े","{":"ि","[":"ि","¥":"ृ","©":"य","¡":"ौ","¨":"ि",
    "A":"अ","B":"आ","C":"उ","D":"ऊ","E":"ए","F":"ऊ","G":"ऋ",
    "§":"ं","|":"ं",":":"ः","`":"र्","«":"र्",
    "$":"","ß":"्",
    "Ð":"द्र","à":"प्र","é":"रु","î":"ष्",
    "Î":"त्त","ñ":"स्","Þ":"न्न","º":"ो","q":"ि","ì":"व्ह",
}

PRE_MATRAS  = {"{","[","¨"}
MATRA_CHARS = {"","ा","ि","ी","ु","ू","े","ो","ौ","ृ","ं","ः"," ","\n","\t"}


def is_unicode_devanagari(text):
    deva  = sum(1 for c in text if "\u0900" <= c <= "\u097F")
    alpha = sum(1 for c in text if c.isascii() and c.isalpha())
    total = deva + alpha
    return (deva / total > 0.3) if total > 0 else False


def deva_digits_to_ascii(text):
    return text.translate(str.maketrans("०१२३४५६७८९","0123456789"))


def convert_shreelipi(text):
    if not text or not isinstance(text, str):
        return str(text) if text else ""
    text = str(text).strip()
    if not text:
        return ""
    if is_unicode_devanagari(text):
        return deva_digits_to_ascii(text)

    result, i, pending = [], 0, None
    while i < len(text):
        matched = False
        for dg, rep in DIGRAPHS:
            if text[i:i+len(dg)] == dg:
                if pending:
                    result.append(rep); result.append(pending); pending = None
                else:
                    result.append(rep)
                i += len(dg); matched = True; break
        if matched:
            continue
        c = text[i]
        if c == "$":
            i += 1; continue
        if c == ">":
            nxt = text[i+1] if i+1 < len(text) else ""
            nxt_m = CHAR_MAP.get(nxt, nxt)
            if nxt and nxt not in (" ","\n") and len(nxt_m)==1 and "\u0915"<=nxt_m<="\u0939":
                result.append("्")
            i += 1; continue
        if c in PRE_MATRAS:
            pending = "ि"; i += 1; continue
        mapped = CHAR_MAP.get(c, c)
        if pending and mapped not in MATRA_CHARS:
            result.append(mapped); result.append(pending); pending = None
        else:
            if pending and c in (" ","\n","\t"):
                result.append(pending); pending = None
            result.append(mapped)
        i += 1
    if pending:
        result.append(pending)
    return "".join(result)


# ─────────────────────────────────────────────────────────
# DEVANAGARI → ENGLISH TRANSLITERATION
# ─────────────────────────────────────────────────────────

DEVA_CONSONANTS = set(range(0x0915, 0x093A))


def _ends_with_bare_consonant(deva_word):
    for c in reversed(deva_word):
        if "\u0900" <= c <= "\u097F":
            return ord(c) in DEVA_CONSONANTS
    return False


def devanagari_to_english(text):
    if not text or not isinstance(text, str):
        return str(text) if text else ""
    text = str(text).strip()
    if not text:
        return ""

    out_words = []
    for word in text.split():
        if not any("\u0900" <= c <= "\u097F" for c in word):
            out_words.append(word)
            continue

        strip_a = _ends_with_bare_consonant(word)
        roman   = transliterate(word, sanscript.DEVANAGARI, sanscript.IAST)

        # Context-sensitive anusvara
        roman = re.sub(r"[ṃṁ]([bpBPmM])", r"m\1", roman)
        roman = re.sub(r"[ṃṁ]([gkGKcCjJ])", r"ng\1", roman)
        roman = re.sub(r"[ṃṁ]", "n", roman)

        for src, dst in [
            ("ā","a"),("Ā","A"),("ī","i"),("Ī","I"),("ū","u"),("Ū","U"),
            ("ṛ","ri"),("Ṛ","Ri"),("ṭ","t"),("Ṭ","T"),("ḍ","d"),("Ḍ","D"),
            ("ṇ","n"),("Ṇ","N"),("ṣ","sh"),("Ṣ","Sh"),("ś","sh"),("Ś","Sh"),
            ("ḥ","h"),("Ḥ","H"),("ñ","ny"),("Ñ","Ny"),
            ("ḷ","l"),("Ḷ","L"),("ḻ","l"),("Ḻ","L"),
        ]:
            roman = roman.replace(src, dst)

        roman = re.sub(r"ngg+","ng", roman)
        roman = re.sub(r"ngk","nk", roman)

        if strip_a and roman.endswith("a"):
            roman = roman[:-1]

        out_words.append(roman.capitalize())

    return " ".join(out_words)


def shreelipi_to_english(text):
    """Full pipeline: Shree Leepi → Devanagari → English."""
    return devanagari_to_english(convert_shreelipi(text))


# ─────────────────────────────────────────────────────────
# MOBILE NUMBER PROCESSING
# ─────────────────────────────────────────────────────────

def extract_mobile_numbers(raw):
    if not raw or not isinstance(raw, str):
        return []
    raw = deva_digits_to_ascii(str(raw).strip())
    if not raw or raw.lower() in ("nan","none",""):
        return []
    parts = re.split(r"[/,;|&\s]+", raw)
    valid = []
    for part in parts:
        digits = re.sub(r"\D","", part.strip())
        if not digits:
            continue
        if len(digits)==12 and digits.startswith("91"):
            digits = digits[2:]
        elif len(digits)==13 and digits.startswith("091"):
            digits = digits[3:]
        if len(digits)==10 and digits[0] in "6789":
            valid.append(digits)
    return valid


# ─────────────────────────────────────────────────────────
# MAIN PIPELINE  (with name-conversion cache)
# ─────────────────────────────────────────────────────────

def process_excel(input_path, output_path):
    t_start = time.time()

    print(f"\n{'='*58}")
    print(f"  Shree Leepi → English Converter  (50,000+ record ready)")
    print(f"{'='*58}")
    print(f"Input : {input_path}")
    print(f"Output: {output_path}\n")

    print("Reading file...", end=" ", flush=True)
    df = pd.read_excel(input_path)
    print(f"{len(df):,} rows loaded ({time.time()-t_start:.1f}s)")

    name_col   = df.columns[0]
    mobile_col = df.columns[1]
    print(f"Columns: '{name_col}', '{mobile_col}'\n")

    # ── Pre-build conversion cache for unique names ───────
    print("Building name conversion cache...", end=" ", flush=True)
    unique_names = df[name_col].dropna().astype(str).str.strip().unique()
    name_cache = {}
    for raw in unique_names:
        if raw and raw.lower() not in ("nan","none",""):
            name_cache[raw] = shreelipi_to_english(raw)
    print(f"{len(name_cache):,} unique names converted ({time.time()-t_start:.1f}s)")

    # ── Process rows ──────────────────────────────────────
    print("Processing rows...", end=" ", flush=True)
    rows_out     = []
    skip_blank   = 0
    skip_invalid = 0
    split_added  = 0

    for _, row in df.iterrows():
        raw_name   = str(row[name_col]).strip()   if pd.notna(row[name_col])   else ""
        raw_mobile = str(row[mobile_col]).strip() if pd.notna(row[mobile_col]) else ""

        if not raw_name or raw_name.lower() in ("nan","none"):
            skip_blank += 1; continue

        name_eng = name_cache.get(raw_name, "")
        if not name_eng.strip():
            skip_blank += 1; continue

        numbers = extract_mobile_numbers(raw_mobile)
        if not numbers:
            skip_invalid += 1; continue

        for num in numbers:
            rows_out.append({"Name": name_eng, "Mobile Number": num})
        if len(numbers) > 1:
            split_added += len(numbers) - 1

    print(f"done ({time.time()-t_start:.1f}s)")

    # ── Deduplicate ───────────────────────────────────────
    df_out  = pd.DataFrame(rows_out, columns=["Name","Mobile Number"])
    before  = len(df_out)
    df_out  = df_out.drop_duplicates()
    deduped = before - len(df_out)

    # ── Summary ───────────────────────────────────────────
    print(f"\nProcessing summary:")
    print(f"  Total input rows             : {len(df):,}")
    print(f"  Skipped (blank)              : {skip_blank:,}")
    print(f"  Skipped (invalid mobile)     : {skip_invalid:,}")
    print(f"  Extra rows from splits       : {split_added:,}")
    print(f"  Duplicates removed           : {deduped:,}")
    print(f"  ─────────────────────────────────")
    print(f"  Final output rows            : {len(df_out):,}")

    # ── Save with formatting ──────────────────────────────
    print(f"\nWriting Excel...", end=" ", flush=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "English Output"

    for ci, h in enumerate(["Name","Mobile Number"], 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color="1A5276")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ri, (_, row) in enumerate(df_out.iterrows(), 2):
        ws.cell(row=ri, column=1, value=row["Name"]).font        = Font(name="Arial", size=11)
        ws.cell(row=ri, column=2, value=row["Mobile Number"]).font = Font(name="Arial", size=11)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"
    wb.save(output_path)

    elapsed = time.time() - t_start
    print(f"done")
    print(f"\n✓ Saved → {output_path}")
    print(f"✓ Total time: {elapsed:.1f} seconds")
    return df_out


# ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    input_file  = sys.argv[1] if len(sys.argv) >= 2 else "/mnt/user-data/uploads/sample_data_to_work.xlsx"
    output_file = sys.argv[2] if len(sys.argv) >= 3 else "/mnt/user-data/outputs/english_output.xlsx"

    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}"); sys.exit(1)

    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    result = process_excel(input_file, output_file)

    print(f"\nSample output (first 15 rows):")
    print(result.head(15).to_string(index=False))
